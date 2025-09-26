#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import re
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Optional, List

import pandas as pd
import yaml


# =========================
# Helpers de normalização
# =========================
def _norm_key(s: str) -> str:
    """Normaliza rótulos: troca quebras de linha por espaço, colapsa espaços,
    remove acentos e põe em minúsculas."""
    if s is None:
        return ""
    s = s.replace("\n", " ").strip()
    s = re.sub(r"\s+", " ", s)
    s = "".join(ch for ch in unicodedata.normalize("NFD", s)
                if unicodedata.category(ch) != "Mn")
    return s.lower()


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    # Limpa os nomes de coluna vindos do source
    df.columns = [c.replace("\n", " ").strip() for c in df.columns]
    df.columns = [re.sub(r"\s+", " ", c) for c in df.columns]
    return df

def _append_rows_to_sheet_preserving_table(path: Path, sheet: str, df: pd.DataFrame):
    """
    Anexa df ao final da aba `sheet`, preservando as demais abas.
    Se a aba possuir Tabela Excel (ListObject), atualiza o range da tabela.
    """
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.exceptions import InvalidFileException
    from zipfile import BadZipFile

    if df is None or df.empty:
        return

    try:
        if not path.exists():
            # cria arquivo novo com header + dados
            with pd.ExcelWriter(path, engine="openpyxl", mode="w") as writer:
                df.to_excel(writer, sheet_name=sheet, index=False)
            return

        # abre workbook existente
        wb = load_workbook(path)
    except (InvalidFileException, BadZipFile, KeyError):
        # arquivo inválido/corrompido → recria do zero
        with pd.ExcelWriter(path, engine="openpyxl", mode="w") as writer:
            df.to_excel(writer, sheet_name=sheet, index=False)
        return

    # cria aba se não existir
    if sheet not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet)
        ws.append(list(df.columns))  # cabeçalho
    else:
        ws = wb[sheet]

    # lê header atual da planilha (1ª linha)
    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [c.value for c in header_cells]

    # reordena/filtra df conforme os headers da planilha
    cols = [c for c in headers if c in df.columns]
    df_to_append = df[cols].copy()

    # anexa linha por linha
    for row in df_to_append.itertuples(index=False, name=None):
        ws.append(list(row))

    # atualiza range das tabelas (se existirem)
    max_row = ws.max_row
    max_col = len(headers)
    if max_col > 0:
        ref = f"A1:{get_column_letter(max_col)}{max_row}"
        for tbl in ws.tables.values():
            tbl.ref = ref

    wb.save(path)


# =========================
# IO
# =========================
def load_config(cfg_path: Path) -> dict:
    with open(cfg_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def read_source(path: Path,
                sheet: Optional[str] = None,
                header_row: Optional[int] = None,
                header_rows: Optional[List[int]] = None) -> pd.DataFrame:
    """
    Lê XLSX/CSV.
    - Se header_rows for passado (duas linhas), usa e achata.
    - Se header_row for passado (uma linha), usa.
    - Caso contrário, AUTODETECT: encontra a linha de cabeçalho sozinho
    e detecta se há uma linha de "blocos" acima (ex.: 'Manifestação', 'Resposta').
    """
    ext = path.suffix.lower()

    # CSV
    if ext == ".csv":
        try:
            return pd.read_csv(path, sep=";", encoding="utf-8")
        except Exception:
            return pd.read_csv(path, sep=";", encoding="latin-1")

    # XLSX/XLS
    if header_rows:
        hdr = [int(r) for r in header_rows]
        df = pd.read_excel(path, sheet_name=sheet, header=hdr)
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = [" ".join([str(x) for x in tup if str(x) != "nan"]).strip()
                        for tup in df.columns.values]
        return df

    if header_row is not None:
        h = int(header_row)  # já é 0-based (pandas)
        return pd.read_excel(path, sheet_name=sheet, header=h)

    # === AUTODETECT ===
    probe = pd.read_excel(path, sheet_name=sheet, header=None, nrows=15)
    candidates = ["situação", "nup", "tipo", "registrado", "assunto", "subassunto",
                "data de abertura", "prazo", "canal de entrada"]

    def score_row(vals):
        vals_norm = [str(v).strip().lower() for v in vals]
        return sum(any(c in v for v in vals_norm) for c in candidates)

    header_idx = None
    for i in range(len(probe)):
        if score_row(probe.iloc[i].values) >= 3:
            header_idx = i
            break
    if header_idx is None:
        header_idx = 3  # chute sensato para exports do Fala.br

    # Detecta se existe linha de "blocos" acima (Manifestação/Resposta)
    use_multi = False
    if header_idx > 0:
        upper_vals = [str(v).strip().lower() for v in list(probe.iloc[header_idx - 1].values)]
        if any(("manifestação" in v) or ("resposta" in v) for v in upper_vals):
            use_multi = True

    if use_multi:
        df = pd.read_excel(path, sheet_name=sheet, header=[header_idx - 1, header_idx])
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = [" ".join([str(x) for x in tup if str(x) != "nan"]).strip()
                        for tup in df.columns.values]
        return df

    return pd.read_excel(path, sheet_name=sheet, header=header_idx)


def read_master(master_file: Path, sheet: str) -> pd.DataFrame:
    if master_file.exists():
        try:
            return pd.read_excel(master_file, sheet_name=sheet)
        except Exception:
            return pd.DataFrame()
    return pd.DataFrame()


def write_master(path: Path, sheet: str, df: pd.DataFrame):
    """
    Substitui apenas a aba `sheet`. Se o arquivo estiver corrompido/inválido,
    recria um novo workbook com a aba `sheet`. Se estiver aberto, salva backup.
    """
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException
    from zipfile import BadZipFile

    try:
        if path.exists():
            try:
                # tenta abrir o workbook existente (só funciona se for xlsx válido)
                wb = load_workbook(path)
                wb.close()
                # se deu certo, substitui só a aba alvo
                with pd.ExcelWriter(
                    path,
                    engine="openpyxl",
                    mode="a",
                    if_sheet_exists="replace",
                ) as writer:
                    df.to_excel(writer, sheet_name=sheet, index=False)
            except (InvalidFileException, BadZipFile, KeyError):
                # arquivo inválido/corrompido -> recria do zero com a aba sheet
                print("⚠️ Arquivo inválido/corrompido. Vou recriar um novo workbook com a aba alvo.")
                with pd.ExcelWriter(path, engine="openpyxl", mode="w") as writer:
                    df.to_excel(writer, sheet_name=sheet, index=False)
        else:
            # cria arquivo do zero
            with pd.ExcelWriter(path, engine="openpyxl", mode="w") as writer:
                df.to_excel(writer, sheet_name=sheet, index=False)
    except PermissionError:
        backup_path = path.with_name(path.stem + "_etl.xlsx")
        with pd.ExcelWriter(backup_path, engine="openpyxl", mode="w") as writer:
            df.to_excel(writer, sheet_name=sheet, index=False)
        print(f"⚠️ Arquivo em uso, salvei no backup: {backup_path}")



# =========================
# Transformações
# =========================
def apply_mapping(df_src: pd.DataFrame, mapping: dict, master_cols: list) -> pd.DataFrame:
    # Mapa "normalizado -> original" das colunas do source
    src_norm_to_orig = {_norm_key(c): c for c in df_src.columns}

    resolved, missing = {}, []
    for src_key, target_col in mapping.items():
        orig = src_norm_to_orig.get(_norm_key(src_key))
        if not orig:
            missing.append(src_key)
        else:
            resolved[orig] = target_col

    if missing:
        available = ", ".join([f"'{c}'" for c in df_src.columns])
        raise KeyError(
            f"As colunas abaixo não foram encontradas no arquivo de origem (após normalização): {missing}\n"
            f"Disponíveis no arquivo: {available}"
        )

    df = df_src[list(resolved.keys())].rename(columns=resolved)

    # Completa colunas do master que faltam (vão como vazio)
    for col in master_cols:
        if col not in df.columns:
            df[col] = pd.NA

    # Reordena conforme master
    df = df[master_cols]
    return df


def parse_dates(df: pd.DataFrame, date_cols: list, out_fmt: str) -> pd.DataFrame:
    # dayfirst=True (Brasil) + preserva vazio como None (evita <NA> no Excel)
    for col in date_cols or []:
        if col in df.columns:
            s = pd.to_datetime(df[col], errors="coerce", dayfirst=True)
            df[col] = s.dt.strftime(out_fmt)
            df[col] = df[col].where(s.notna(), None)
    return df


def upsert(master: pd.DataFrame, incoming: pd.DataFrame, key: str):
    if master.empty:
        new_rows = incoming.copy()
        updated_rows = pd.DataFrame(columns=incoming.columns)
        return incoming.copy(), new_rows, updated_rows

    if key not in master.columns:
        raise KeyError(f"A coluna de chave única '{key}' não existe no master.")

    master = master.copy()
    master_keys = set(master[key].astype(str).fillna(""))
    incoming_keys = set(incoming[key].astype(str).fillna(""))

    new_mask = ~incoming[key].astype(str).isin(master[key].astype(str))
    new_rows = incoming[new_mask].copy()

    existing_keys = incoming_keys.intersection(master_keys)
    updated_rows = pd.DataFrame(columns=incoming.columns)

    if existing_keys:
        idx_map = {str(k): i for i, k in enumerate(master[key].astype(str))}
        for _, row in incoming[~new_mask].iterrows():
            k = str(row[key])
            if k in idx_map:
                i = idx_map[k]
                for col in incoming.columns:
                    master.at[i, col] = row[col]
        updated_rows = incoming[~new_mask].copy()

    result = pd.concat([master, new_rows], ignore_index=True)
    return result, new_rows, updated_rows


def to_excel_safe(df: pd.DataFrame) -> pd.DataFrame:
    # Troca pd.NA/NaN/NaT por None (compatível com openpyxl)
    return df.astype(object).where(pd.notnull(df), None)


# =========================
# Main
# =========================
cfg = load_config("config.yaml")
opts = cfg.get("options", {})   # <--- adiciona isso
append_only = bool(opts.get("append_only", False))


def main():
    cfg = load_config(Path("config.yaml"))

    # Paths e opções
    src_path = Path(cfg["paths"]["source_file"])
    master_path = Path(cfg["paths"]["master_file"])
    master_sheet = cfg["paths"].get("master_sheet", "BD")
    out_dir = Path(cfg["paths"].get("out_dir", "./etl_out"))
    out_dir.mkdir(parents=True, exist_ok=True)

    source_sheet = cfg["paths"].get("source_sheet")
    source_header_row = cfg["paths"].get("source_header_row")  # pode ser None (autodetect)
    source_header_rows = cfg["paths"].get("source_header_rows")  # pode ser None (autodetect)

    mapping = cfg["mapping"]
    opts = cfg["options"]
    unique_key = opts.get("unique_key", "NUP")
    date_cols = opts.get("date_cols", [])
    out_fmt = opts.get("output_date_format", "%Y-%m-%d")
    upsert_on_conflict = bool(opts.get("upsert_on_conflict", True))
    write_change_sets = bool(opts.get("write_change_sets", True))

    # 1) Ler origem
    df_src = read_source(src_path, sheet=source_sheet,
                        header_row=source_header_row,
                        header_rows=source_header_rows)
    df_src = normalize_columns(df_src)
    print("Colunas lidas:", df_src.columns.tolist())

    # 2) Ler master atual (se existir)
    df_master_existing = read_master(master_path, master_sheet)

    # 3) Descobrir colunas do master (ordem final)
    if not df_master_existing.empty:
        master_cols = list(df_master_existing.columns)
        for c in mapping.values():
            if c not in master_cols:
                master_cols.append(c)
    else:
        master_cols = list(dict.fromkeys(list(mapping.values()) + ["DataImportacao", "Mês"]))

    # 4) Aplicar mapping (origem -> nomes do master)
    df_in = apply_mapping(df_src, mapping, master_cols)

    # 5) DataImportacao
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if "DataImportacao" in df_in.columns:
        df_in["DataImportacao"] = now_str

    # 6) Datas
    df_in = parse_dates(df_in, date_cols, out_fmt)

    # 7) Mês a partir de Data de Abertura (se existir no master)
    if "Mês" in master_cols and "Data de Abertura" in df_in.columns:
        tmp_dt = pd.to_datetime(df_in["Data de Abertura"], errors="coerce", dayfirst=True)
        df_in["Mês"] = tmp_dt.dt.strftime("%m/%Y")

    # 8) Preparar master atual com mesmas colunas/ordem
    master_cur = df_master_existing if not df_master_existing.empty else pd.DataFrame(columns=master_cols)
    for c in master_cols:
        if c not in master_cur.columns:
            master_cur[c] = pd.NA
    master_cur = master_cur[master_cols]
    df_in = df_in[master_cols]

    # 9) Upsert
    if unique_key not in df_in.columns:
        raise KeyError(f"A coluna de chave única '{unique_key}' não existe após o mapeamento.")
    merged, new_rows, updated_rows = upsert(master_cur, df_in, unique_key)

    # 10) CSVs de novos/atualizados e log
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    if write_change_sets:
        if not new_rows.empty:
            new_rows.to_csv(out_dir / f"novos_{ts}.csv", index=False, encoding="utf-8-sig", sep=";")
        if not updated_rows.empty:
            updated_rows.to_csv(out_dir / f"atualizados_{ts}.csv", index=False, encoding="utf-8-sig", sep=";")
    summary = {
        "timestamp": ts,
        "source_file": str(src_path),
        "master_file": str(master_path),
        "master_sheet": master_sheet,
        "rows_source": int(df_in.shape[0]),
        "rows_master_before": int(master_cur.shape[0]),
        "rows_master_after": int(merged.shape[0]),
        "new_rows": int(new_rows.shape[0]),
        "updated_rows": int(updated_rows.shape[0]),
    }
    with open(out_dir / f"log_{ts}.json", "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

# 11) Salvar no Excel
append_only = bool(opts.get("append_only", False))

if append_only:
    # Só anexar os NOVOS no final da aba (ignora atualizações)
    df_to_append = to_excel_safe(new_rows)
    _append_rows_to_sheet_preserving_table(master_path, master_sheet, df_to_append)
else:
    # Comportamento normal: upsert completo e substitui apenas a aba alvo
    merged = to_excel_safe(merged)
    write_master(master_path, master_sheet, merged)





if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print("❌ Erro na execução do ETL:", e)
        raise
